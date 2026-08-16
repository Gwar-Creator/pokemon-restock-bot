import json
import math
import os
import statistics
import tempfile
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

API = "https://tcg-api-production-5148.up.railway.app"
KEY = os.getenv("TCG_CARDMARKET_API_KEY", "").strip()
WEBHOOK = (
    os.getenv("CARDMARKET_WEBHOOK_URL", "").strip()
    or os.getenv("PRICE_HISTORY_WEBHOOK_URL", "").strip()
)
TZ = os.getenv("CARDMARKET_TIMEZONE", "Europe/Copenhagen")
HOUR = int(os.getenv("CARDMARKET_DAILY_HOUR", "8") or 8)
FORCE = os.getenv("CARDMARKET_FORCE_RUN", "0") == "1"
STATE = Path("cardmarket_chase_state.json")
WATCH = Path("cardmarket_chase_watchlist.json")
RESTOCK_STATE = Path("restock_state_v2.json")
SOURCE = "https://www.tcg-cardmarket-api.com/"
HISTORY_DAYS = 35

GAME_COLOR = {"POKÉMON": 0x5865F2, "LORCANA": 0x9B59B6}
GAME_ICON = {"POKÉMON": "⚡", "LORCANA": "✨"}
OPPORTUNITY_TYPES = {"BOOSTER BOX", "ETB", "BOOSTER BUNDLE"}
FOREIGN_MARKERS = {
    "japansk", "japanese", "kinesisk", "chinese", "korean", "koreansk",
    "german", "tysk", "french", "fransk", "spanish", "spansk", "italian",
}


def f(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clamp(value, low=0.0, high=100.0):
    return max(low, min(high, value))


def pct(current, baseline):
    current, baseline = f(current), f(baseline)
    if current is None or baseline is None or baseline <= 0:
        return None
    return (current / baseline - 1.0) * 100.0


def eur(value):
    value = f(value)
    if value is None:
        return "–"
    if value >= 1000:
        return f"€{value:,.0f}".replace(",", ".")
    return f"€{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def dkk(value):
    value = f(value)
    if value is None:
        return "–"
    if abs(value - round(value)) < 0.005:
        return f"{int(round(value)):,} kr.".replace(",", ".")
    return f"{value:,.2f} kr.".replace(",", "X").replace(".", ",").replace("X", ".")


def pc(value):
    value = f(value)
    if value is None:
        return "–"
    sign = "+" if value > 0.05 else ""
    return f"{sign}{value:.1f}%".replace(".", ",")


def load(path, default):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def save(path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def post(embeds=None, content=None, file=None):
    if not WEBHOOK:
        return False

    payload = {
        "username": "MasterBot",
        "allowed_mentions": {"parse": []},
    }
    if embeds:
        payload["embeds"] = embeds[:10]
    if content:
        payload["content"] = content[:2000]

    if file:
        with open(file, "rb") as handle:
            response = requests.post(
                WEBHOOK,
                data={"payload_json": json.dumps(payload, ensure_ascii=False)},
                files={
                    "files[0]": (
                        Path(file).name,
                        handle,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                timeout=45,
            )
    else:
        response = requests.post(WEBHOOK, json=payload, timeout=30)

    response.raise_for_status()
    return True


def emb(title, text, color, footer=None):
    result = {
        "title": title[:256],
        "description": (text or " ")[:4096],
        "color": color,
    }
    if footer:
        result["footer"] = {"text": footer[:2048]}
    return result


def normalize_text(value):
    value = unicodedata.normalize("NFKD", str(value or "").lower())
    value = "".join(char for char in value if not unicodedata.combining(char))
    cleaned = []
    for char in value:
        cleaned.append(char if char.isalnum() else " ")
    return " ".join("".join(cleaned).split())


def score_bar(score):
    score = int(round(clamp(f(score) or 0)))
    filled = int(round(score / 10.0))
    return "█" * filled + "░" * (10 - filled)


def heat_label(score):
    score = f(score)
    if score is None:
        return "NO DATA"
    if score >= 75:
        return "HOT"
    if score >= 60:
        return "WARM"
    if score >= 40:
        return "NEUTRAL"
    if score >= 25:
        return "COOL"
    return "COLD"


def heat_icon(score):
    score = f(score) or 0
    if score >= 75:
        return "🔥"
    if score >= 60:
        return "🟠"
    if score >= 40:
        return "🟡"
    if score >= 25:
        return "🔵"
    return "🧊"


def scale_centered(value, span=12.0):
    value = f(value)
    if value is None:
        return None
    return clamp(50.0 + (value / span) * 50.0)


def fetch_cards(watch):
    session = requests.Session()
    session.headers.update(
        {
            "X-API-Key": KEY,
            "Accept": "application/json",
            "User-Agent": "Pokemon-Lorcana-MasterBot/1.6",
        }
    )

    output = []
    used = 0
    remaining = None
    limit = None

    for game in watch["games"]:
        for set_info in game["sets"]:
            card_ids = [str(value) for value in set_info["cardIds"]]
            for index in range(0, len(card_ids), 10):
                response = session.post(
                    API + "/cards/batch",
                    json={"game": game["slug"], "cardIds": card_ids[index:index + 10]},
                    timeout=30,
                )
                used += 1

                try:
                    remaining = int(response.headers.get("X-RateLimit-Remaining", ""))
                    limit = int(response.headers.get("X-RateLimit-Limit", ""))
                except (TypeError, ValueError):
                    pass

                if response.status_code == 429:
                    raise RuntimeError("TCG Cardmarket API rate limit ramt")

                response.raise_for_status()
                rows = response.json()
                if isinstance(rows, dict):
                    rows = rows.get("data", rows)
                if not isinstance(rows, list):
                    raise RuntimeError("Uventet Cardmarket API svar")

                for raw in rows:
                    prices = raw.get("price") or {}
                    trend = f(prices.get("trend"))
                    foil_trend = f(prices.get("foilTrend"))
                    low = f(prices.get("low"))
                    foil_low = f(prices.get("foilLow"))

                    market = trend
                    variant = "Normal"
                    if foil_trend is not None and (market is None or foil_trend > market):
                        market, variant = foil_trend, "Foil"
                    if market is None:
                        market = low
                    if market is None and foil_low is not None:
                        market, variant = foil_low, "Foil"

                    output.append(
                        {
                            "game": game["game"],
                            "set": set_info["name"],
                            "expansion": set_info["expansionId"],
                            "id": str(raw.get("externalId")),
                            "name": raw.get("name") or "Ukendt kort",
                            "variant": variant,
                            "market": market,
                            "low": low,
                            "trend": trend,
                            "avg1": f(prices.get("avg1")),
                            "avg7": f(prices.get("avg7")),
                            "avg30": f(prices.get("avg30")),
                            "foilLow": foil_low,
                            "foilTrend": foil_trend,
                            "updated": prices.get("updatedAt") or "",
                        }
                    )

    return output, used, remaining, limit


def history_change(history, days, current_market, current_day):
    if not history or current_market is None:
        return None

    target = current_day - timedelta(days=days)
    candidates = []
    for row in history:
        try:
            row_day = date.fromisoformat(str(row.get("date")))
        except (TypeError, ValueError):
            continue
        if row_day <= target and f(row.get("market")) is not None:
            candidates.append((row_day, f(row.get("market"))))

    if not candidates:
        return None

    _, baseline = max(candidates, key=lambda item: item[0])
    return pct(current_market, baseline)


def add_history(cards, old, stamp, current_day):
    previous = old.get("cards", {}) if isinstance(old, dict) else {}
    next_cards = {}
    lows = []

    for card in cards:
        key = card["game"] + "|" + card["id"]
        prior = previous.get(key, {}) if isinstance(previous.get(key, {}), dict) else {}

        card["daily"] = pct(card["market"], prior.get("market"))
        old_low = f(prior.get("histLow"))
        current_low = card["low"]
        card["histLow"] = (
            current_low
            if old_low is None
            else min(old_low, current_low) if current_low is not None else old_low
        )
        card["histLowAt"] = (
            stamp
            if old_low is None or (current_low is not None and current_low < old_low)
            else prior.get("histLowAt", stamp)
        )

        if (
            old_low
            and current_low is not None
            and current_low <= old_low * 0.98
            and (card["market"] or 0) >= 5
        ):
            lows.append({**card, "oldLow": old_low})

        history = prior.get("history", [])
        if not isinstance(history, list):
            history = []
        history = [row for row in history if isinstance(row, dict)]
        today_text = current_day.isoformat()
        history = [row for row in history if str(row.get("date")) != today_text]
        history.append(
            {
                "date": today_text,
                "market": card.get("market"),
                "low": card.get("low"),
            }
        )
        history = history[-HISTORY_DAYS:]

        card["history"] = history
        card["weekly"] = history_change(history, 7, card.get("market"), current_day)
        card["firstSeen"] = prior.get("firstSeen", stamp)
        card["lastSeen"] = stamp
        next_cards[key] = dict(card)

    return next_cards, lows


def ranked(cards):
    groups = {}
    for card in cards:
        groups.setdefault((card["game"], card["set"]), []).append(card)

    output = []
    for rows in groups.values():
        rows = sorted(rows, key=lambda row: f(row["market"]) or -1, reverse=True)
        for index, card in enumerate(rows, 1):
            output.append({**card, "rank": index})
    return output


def heat_metrics(rows):
    momentum = [
        value
        for value in (pct(card.get("avg7"), card.get("avg30")) for card in rows)
        if value is not None
    ]
    acceleration = [
        value
        for value in (pct(card.get("avg1"), card.get("avg7")) for card in rows)
        if value is not None
    ]

    ordered = sorted(rows, key=lambda row: f(row.get("market")) or -1, reverse=True)
    top3 = ordered[:3]
    top3_momentum = [
        value
        for value in (pct(card.get("avg7"), card.get("avg30")) for card in top3)
        if value is not None
    ]

    median_momentum = statistics.median(momentum) if momentum else None
    median_acceleration = statistics.median(acceleration) if acceleration else None
    median_top3 = statistics.median(top3_momentum) if top3_momentum else None
    rising_share = (
        sum(1 for value in momentum if value > 0) / len(momentum) * 100.0
        if momentum
        else None
    )

    components = []
    for value, weight in (
        (scale_centered(median_momentum), 0.45),
        (rising_share, 0.25),
        (scale_centered(median_top3), 0.20),
        (scale_centered(median_acceleration), 0.10),
    ):
        if value is not None:
            components.append((value, weight))

    if components:
        total_weight = sum(weight for _, weight in components)
        score = sum(value * weight for value, weight in components) / total_weight
    else:
        score = 50.0

    return {
        "score": round(clamp(score), 1),
        "label": heat_label(score),
        "momentum": median_momentum,
        "acceleration": median_acceleration,
        "top3_momentum": median_top3,
        "rising_share": rising_share,
        "rising_count": sum(1 for value in momentum if value > 0),
        "measured_count": len(momentum),
    }


def build_set_summaries(cards, game=None, old_scores=None):
    old_scores = old_scores or {}
    groups = {}
    for card in cards:
        if game and card["game"] != game:
            continue
        groups.setdefault((card["game"], card["set"]), []).append(card)

    output = []
    for (game_name, set_name), rows in groups.items():
        metrics = heat_metrics(rows)
        top = max(rows, key=lambda row: f(row.get("market")) or -1)
        score_key = game_name + "|" + set_name
        previous_score = f(old_scores.get(score_key))
        score_delta = None if previous_score is None else metrics["score"] - previous_score
        output.append(
            {
                "game": game_name,
                "set": set_name,
                "cards": len(rows),
                "top": top,
                "score_delta": score_delta,
                **metrics,
            }
        )

    return output


def set_summary_map(summaries):
    return {(row["game"], row["set"]): row for row in summaries}


def price_history_products():
    state = load(RESTOCK_STATE, {})
    history = state.get("price_history", {}) if isinstance(state, dict) else {}
    products = history.get("products", {}) if isinstance(history, dict) else {}
    return products if isinstance(products, dict) else {}


def product_parts(product_key):
    parts = str(product_key).split("|", 3)
    if len(parts) != 4:
        return None
    return {
        "game": parts[0],
        "type": parts[1],
        "language": parts[2],
        "label": parts[3],
    }


def set_matches_product(set_name, label, entry_name):
    wanted = normalize_text(set_name)
    candidate = normalize_text(label + " " + str(entry_name or ""))
    if not wanted or not candidate:
        return False
    if any(marker in candidate.split() for marker in FOREIGN_MARKERS):
        return False
    return wanted in candidate


def sealed_opportunities(summaries, products):
    opportunities = []

    for summary in summaries:
        candidates = []
        for product_key, entry in products.items():
            if not isinstance(entry, dict):
                continue
            parts = product_parts(product_key)
            if not parts:
                continue
            if parts["game"] != summary["game"]:
                continue
            if parts["type"] not in OPPORTUNITY_TYPES:
                continue
            if not set_matches_product(summary["set"], parts["label"], entry.get("name")):
                continue

            current = f(entry.get("current_best"))
            historical = f(entry.get("historical_low"))
            if current is None or historical is None or historical <= 0:
                continue

            gap = max(0.0, pct(current, historical) or 0.0)
            price_score = clamp(100.0 - gap * 5.0)
            opportunity_score = summary["score"] * 0.68 + price_score * 0.32
            candidates.append(
                {
                    "game": summary["game"],
                    "set": summary["set"],
                    "type": parts["type"],
                    "product": entry.get("name") or parts["label"].title(),
                    "current": current,
                    "historical": historical,
                    "gap": gap,
                    "shop": entry.get("current_shop") or "–",
                    "url": entry.get("current_url") or "",
                    "heat_score": summary["score"],
                    "heat_label": summary["label"],
                    "opportunity_score": round(opportunity_score, 1),
                }
            )

        if candidates:
            best = max(
                candidates,
                key=lambda row: (
                    row["opportunity_score"],
                    -row["gap"],
                    row["current"],
                ),
            )
            opportunities.append(best)

    return sorted(opportunities, key=lambda row: row["opportunity_score"], reverse=True)


def set_block(summary, index=None):
    prefix = f"{index}. " if index else ""
    breadth = (
        f"{summary['rising_count']}/{summary['measured_count']} rising"
        if summary["measured_count"]
        else "ingen 7d/30d data"
    )
    delta = ""
    if summary.get("score_delta") is not None and abs(summary["score_delta"]) >= 0.5:
        delta = f" · Δ {pc(summary['score_delta'])} score"

    return (
        f"**{prefix}{summary['set']}**\n"
        f"{heat_icon(summary['score'])} `{score_bar(summary['score'])}` "
        f"**{summary['score']:.0f}/100 · {summary['label']}**{delta}\n"
        f"7d/30d **{pc(summary['momentum'])}** · {breadth} · Top 3 **{pc(summary['top3_momentum'])}**"
    )


def card_block(card, index=None, extra=None):
    prefix = f"{index}. " if index else ""
    result = (
        f"**{prefix}{card['name']}**\n"
        f"{card['set']}\n"
        f"Trend **{eur(card['market'])}** · Low {eur(card['low'])}"
    )
    if card.get("variant") == "Foil":
        result += " · Foil signal"
    if extra:
        result += "\n" + extra
    return result


def opportunity_block(row, index=None):
    prefix = f"{index}. " if index else ""
    gap_icon = "🟢" if row["gap"] <= 5 else "🟡" if row["gap"] <= 10 else "🟠"
    return (
        f"**{prefix}{row['set']} · {row['type'].title()}**\n"
        f"🇩🇰 Nu **{dkk(row['current'])}** · {row['shop']}\n"
        f"{gap_icon} DK low {dkk(row['historical'])} · **{pc(row['gap'])} over low**\n"
        f"{heat_icon(row['heat_score'])} Singles heat **{row['heat_score']:.0f}/100** · "
        f"💎 Signal **{row['opportunity_score']:.0f}/100**"
    )


def quiet_summary(game, rows, summaries, used):
    color = GAME_COLOR[game]
    icon = GAME_ICON[game]
    hottest = sorted(summaries, key=lambda row: row["score"], reverse=True)[:3]
    text = (
        f"**{len({row['set'] for row in rows})} sæt** · **{len(rows)} chase cards** · "
        f"**{used} API requests i alt**\n\n"
        "✅ Ingen større prisbevægelser, nye lows eller stærke opportunity-signaler i dag.\n\n"
        "**Hottest lige nu**\n\n"
        + "\n\n".join(set_block(row, index) for index, row in enumerate(hottest, 1))
        + "\n\n*Fuld Top 20 pr. sæt er opdateret i Excel.*"
    )
    post(
        embeds=[
            emb(
                f"{icon} {game.title()} · MARKET PULSE · ROLIG DAG",
                text,
                color,
                footer="MasterBot · Card Market Watch",
            )
        ]
    )


def discord_market_pulse(cards, game, summaries, opportunities, first_scores, lows, used, weekly=False):
    color = GAME_COLOR[game]
    icon = GAME_ICON[game]
    rows = [card for card in cards if card["game"] == game]
    game_summaries = [row for row in summaries if row["game"] == game]
    game_opportunities = [row for row in opportunities if row["game"] == game]
    game_lows = [card for card in lows if card["game"] == game]

    movers = [
        card
        for card in rows
        if card.get("daily") is not None
        and abs(card["daily"]) >= 8
        and (f(card.get("market")) or 0) >= 5
    ]
    heat_moves = [
        row
        for row in game_summaries
        if row.get("score_delta") is not None and abs(row["score_delta"]) >= 7
    ]
    strong_opportunities = [row for row in game_opportunities if row["opportunity_score"] >= 72]

    meaningful = bool(first_scores or weekly or game_lows or movers or heat_moves or strong_opportunities)
    if not meaningful:
        quiet_summary(game, rows, game_summaries, used)
        return

    hot_count = sum(1 for row in game_summaries if row["score"] >= 75)
    header = (
        f"**{len(game_summaries)} sæt** · **{len(rows)} chase cards**\n\n"
        f"🔥 HOT sets: **{hot_count}**\n"
        f"💎 Stærke sealed-signaler: **{len(strong_opportunities)}**\n"
        f"🏷️ Nye observerede lows: **{len(game_lows)}**\n"
        f"📊 API: **{used} requests i alt**\n\n"
        "*Heat Score kombinerer 7d/30d momentum, breadth, Top-3 momentum og kortsigtet acceleration.*"
    )

    embeds = [
        emb(
            f"{icon} {game.title()} · MARKET PULSE",
            header,
            color,
            footer="MasterBot · Card Market Watch",
        )
    ]

    hottest = sorted(game_summaries, key=lambda row: row["score"], reverse=True)[:6]
    embeds.append(
        emb(
            "🔥 Hottest sets",
            "\n\n".join(set_block(row, index) for index, row in enumerate(hottest, 1)),
            0xF1C40F,
        )
    )

    if game_opportunities:
        display = game_opportunities[:5]
        embeds.append(
            emb(
                "💎 Sealed + singles · interessante setups",
                "\n\n".join(
                    opportunity_block(row, index) for index, row in enumerate(display, 1)
                )
                + "\n\n*Markedssignal – ikke beregnet opening-EV.*",
                0x57F287,
            )
        )

    top = sorted(rows, key=lambda row: f(row.get("market")) or -1, reverse=True)[:5]
    embeds.append(
        emb(
            "🏆 Top 5 chases lige nu",
            "\n\n".join(card_block(card, index) for index, card in enumerate(top, 1)),
            color,
        )
    )

    if not first_scores:
        up = sorted(
            [card for card in rows if card.get("daily") is not None and card["daily"] >= 5],
            key=lambda row: row["daily"],
            reverse=True,
        )[:5]
        down = sorted(
            [card for card in rows if card.get("daily") is not None and card["daily"] <= -5],
            key=lambda row: row["daily"],
        )[:5]

        if up:
            embeds.append(
                emb(
                    "📈 Risers siden i går",
                    "\n\n".join(
                        card_block(card, index, f"Siden i går **{pc(card['daily'])}**")
                        for index, card in enumerate(up, 1)
                    ),
                    0x57F287,
                )
            )

        if down:
            embeds.append(
                emb(
                    "📉 Fallers siden i går",
                    "\n\n".join(
                        card_block(card, index, f"Siden i går **{pc(card['daily'])}**")
                        for index, card in enumerate(down, 1)
                    ),
                    0xED4245,
                )
            )

        if game_lows:
            embeds.append(
                emb(
                    "🏷️ Nye observerede Cardmarket lows",
                    "\n\n".join(
                        card_block(
                            card,
                            index,
                            f"Low {eur(card['oldLow'])} → **{eur(card['low'])}**",
                        )
                        for index, card in enumerate(game_lows[:5], 1)
                    ),
                    0xF1C40F,
                )
            )

    post(embeds=embeds)


def weekly_card_movers(cards, game):
    rows = [
        card
        for card in cards
        if card["game"] == game and card.get("weekly") is not None and (f(card.get("market")) or 0) >= 5
    ]
    return (
        sorted(rows, key=lambda row: row["weekly"], reverse=True)[:5],
        sorted(rows, key=lambda row: row["weekly"])[:5],
    )


def discord_weekly_report(cards, game, summaries, opportunities):
    color = GAME_COLOR[game]
    icon = GAME_ICON[game]
    game_summaries = [row for row in summaries if row["game"] == game]
    game_opportunities = [row for row in opportunities if row["game"] == game]
    hottest = sorted(game_summaries, key=lambda row: row["score"], reverse=True)[:5]
    coldest = sorted(game_summaries, key=lambda row: row["score"])[:5]
    up, down = weekly_card_movers(cards, game)

    embeds = [
        emb(
            f"🗓️ {icon} {game.title()} · UGENS MARKET REPORT",
            "Ugens samlede snapshot af **set heat, chase-bevægelser og sealed opportunities**.\n\n"
            "*7d-kortbevægelser kræver mindst ca. en uges egen historik og bliver derfor bedre uge for uge.*",
            color,
        ),
        emb(
            "🔥 Ugens hottest sets",
            "\n\n".join(set_block(row, index) for index, row in enumerate(hottest, 1)),
            0xF1C40F,
        ),
        emb(
            "🧊 Coldest sets",
            "\n\n".join(set_block(row, index) for index, row in enumerate(coldest, 1)),
            0x3498DB,
        ),
    ]

    if up:
        embeds.append(
            emb(
                "🚀 Største 7d risers",
                "\n\n".join(
                    card_block(card, index, f"7d **{pc(card['weekly'])}**")
                    for index, card in enumerate(up, 1)
                ),
                0x57F287,
            )
        )
    if down:
        embeds.append(
            emb(
                "🪂 Største 7d fallers",
                "\n\n".join(
                    card_block(card, index, f"7d **{pc(card['weekly'])}**")
                    for index, card in enumerate(down, 1)
                ),
                0xED4245,
            )
        )
    if game_opportunities:
        embeds.append(
            emb(
                "💎 Bedste sealed + singles setups",
                "\n\n".join(
                    opportunity_block(row, index)
                    for index, row in enumerate(game_opportunities[:5], 1)
                ),
                0x57F287,
            )
        )

    post(embeds=embeds)


def workbook(cards, summaries, opportunities, stamp, used, remaining):
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.sheet_view.showGridLines = False

    info = [
        ("CARD MARKET WATCH", "Top 20 chase tracker for Pokémon and Disney Lorcana"),
        ("Updated", stamp),
        ("Source", SOURCE),
        ("API requests", used),
        ("Rate remaining", remaining),
        ("Selection", "20 recent mainstream Pokémon sets + all 13 mainstream Lorcana booster sets"),
        ("Heat Score", "0-100 score using 7d/30d momentum, breadth, Top-3 momentum and 1d/7d acceleration."),
        ("Opportunity", "Combines Heat Score with current Danish sealed price distance from this bot's observed historical low."),
        ("Observed Low", "Lowest API Low observed by this bot; shipping excluded."),
        ("History", f"Stores up to {HISTORY_DAYS} daily observations per tracked card."),
        ("Top 20 method", "Seeded from Cardmarket snapshot 2026-08-07; reranked daily inside the tracked 20. Discovery outside the seed is not automated yet."),
        ("Limitation", "Aggregate Cardmarket data only; no seller country/condition/sales filters."),
    ]
    for row_index, (label, value) in enumerate(info, 1):
        readme.cell(row_index, 1, label).font = Font(bold=True)
        readme.cell(row_index, 2, value).alignment = Alignment(wrap_text=True)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 110

    ranked_cards = ranked(cards)
    summary_lookup = set_summary_map(summaries)

    pulse = workbook.create_sheet("Market Pulse")
    pulse.sheet_view.showGridLines = False
    pulse.append(["Section", "Game", "Set / Card", "Score / Change", "Market", "DK Current", "DK Low", "Shop"])

    for game in ("POKÉMON", "LORCANA"):
        for summary in sorted(
            [row for row in summaries if row["game"] == game],
            key=lambda row: row["score"],
            reverse=True,
        )[:8]:
            pulse.append(["HOT SET", game, summary["set"], summary["score"], None, None, None, None])
        for row in [row for row in opportunities if row["game"] == game][:8]:
            pulse.append([
                "OPPORTUNITY",
                game,
                row["set"] + " · " + row["type"].title(),
                row["opportunity_score"],
                None,
                row["current"],
                row["historical"],
                row["shop"],
            ])
        for card in sorted(
            [row for row in ranked_cards if row["game"] == game],
            key=lambda row: f(row.get("market")) or -1,
            reverse=True,
        )[:10]:
            pulse.append(["TOP CHASE", game, card["name"] + " · " + card["set"], None, card["market"], None, None, None])

    for cell in pulse[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    pulse.freeze_panes = "A2"
    pulse.auto_filter.ref = pulse.dimensions
    for cell in pulse["E"][1:]:
        cell.number_format = '€#,##0.00'
    for column in ("F", "G"):
        for cell in pulse[column][1:]:
            cell.number_format = '#,##0.00 "kr."'
    for column in pulse.columns:
        pulse.column_dimensions[get_column_letter(column[0].column)].width = min(
            55, max(12, max(len(str(cell.value or "")) for cell in column) + 2)
        )

    for game, title, color in (
        ("POKÉMON", "Pokemon Top 20", "17365D"),
        ("LORCANA", "Lorcana Top 20", "7030A0"),
    ):
        sheet = workbook.create_sheet(title)
        sheet.sheet_view.showGridLines = False
        headers = [
            "Rank", "Set", "Card", "CM ID", "Signal", "Market €", "Low €", "Trend €",
            "Avg 1d €", "Avg 7d €", "Avg 30d €", "Foil Low €", "Foil Trend €",
            "Daily change %", "7d own change %", "Observed Low €", "7d vs 30d %",
            "Set Heat Score", "Heat label", "Updated", "Source",
        ]
        sheet.append(headers)

        for card in sorted(
            [row for row in ranked_cards if row["game"] == game],
            key=lambda row: (row["set"], row["rank"]),
        ):
            summary = summary_lookup[(card["game"], card["set"])]
            sheet.append(
                [
                    card["rank"], card["set"], card["name"], card["id"], card["variant"],
                    card["market"], card["low"], card["trend"], card["avg1"], card["avg7"],
                    card["avg30"], card["foilLow"], card["foilTrend"],
                    None if card.get("daily") is None else card["daily"] / 100,
                    None if card.get("weekly") is None else card["weekly"] / 100,
                    card["histLow"],
                    None if pct(card["avg7"], card["avg30"]) is None else pct(card["avg7"], card["avg30"]) / 100,
                    summary["score"], summary["label"], card["updated"], SOURCE,
                ]
            )

        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(color="FFFFFF", bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in "FGHIJKLMNP":
            for cell in sheet[column][1:]:
                cell.number_format = '€#,##0.00'
                cell.font = Font(color="008000")
        for column in ("N", "O", "Q"):
            for cell in sheet[column][1:]:
                cell.number_format = "0.0%"
        for column in sheet.columns:
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(
                52, max(10, max(len(str(cell.value or "")) for cell in column) + 2)
            )
        sheet.column_dimensions["C"].width = 52

    overview = workbook.create_sheet("Set Overview")
    overview_headers = [
        "Game", "Set", "Cards", "Heat Score", "Heat", "Median 7d vs 30d %",
        "Rising %", "Top 3 momentum %", "Top Chase", "Top Chase €",
        "DK Product", "DK Current", "DK Historical Low", "Gap %", "Shop",
        "Opportunity Score", "URL",
    ]
    overview.append(overview_headers)
    opportunity_lookup = {(row["game"], row["set"]): row for row in opportunities}

    for summary in sorted(summaries, key=lambda row: (row["game"], -row["score"], row["set"])):
        opportunity = opportunity_lookup.get((summary["game"], summary["set"]), {})
        overview.append(
            [
                summary["game"], summary["set"], summary["cards"], summary["score"], summary["label"],
                None if summary["momentum"] is None else summary["momentum"] / 100,
                None if summary["rising_share"] is None else summary["rising_share"] / 100,
                None if summary["top3_momentum"] is None else summary["top3_momentum"] / 100,
                summary["top"]["name"], summary["top"]["market"], opportunity.get("product"),
                opportunity.get("current"), opportunity.get("historical"),
                None if opportunity.get("gap") is None else opportunity.get("gap") / 100,
                opportunity.get("shop"), opportunity.get("opportunity_score"), opportunity.get("url"),
            ]
        )

    for cell in overview[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    overview.freeze_panes = "A2"
    overview.auto_filter.ref = overview.dimensions
    for column in ("F", "G", "H", "N"):
        for cell in overview[column][1:]:
            cell.number_format = "0.0%"
    for cell in overview["J"][1:]:
        cell.number_format = '€#,##0.00'
    for column in ("L", "M"):
        for cell in overview[column][1:]:
            cell.number_format = '#,##0.00 "kr."'
    for column in overview.columns:
        overview.column_dimensions[get_column_letter(column[0].column)].width = min(
            55, max(10, max(len(str(cell.value or "")) for cell in column) + 2)
        )
    overview.column_dimensions["Q"].width = 50

    path = Path(tempfile.gettempdir()) / f"cardmarket_top20_{stamp[:10]}.xlsx"
    workbook.save(path)
    return path, ranked_cards


def update_set_history(old, summaries, current_day):
    history = old.get("set_history", {}) if isinstance(old, dict) else {}
    if not isinstance(history, dict):
        history = {}

    history[current_day.isoformat()] = {
        row["game"] + "|" + row["set"]: row["score"] for row in summaries
    }

    cutoff = current_day - timedelta(days=HISTORY_DAYS)
    cleaned = {}
    for day_text, scores in history.items():
        try:
            parsed = date.fromisoformat(day_text)
        except (TypeError, ValueError):
            continue
        if parsed >= cutoff and isinstance(scores, dict):
            cleaned[day_text] = scores
    return cleaned


def current_score_state(summaries):
    return {row["game"] + "|" + row["set"]: row["score"] for row in summaries}


def main():
    if not KEY:
        print("CARD MARKET: TCG_CARDMARKET_API_KEY mangler")
        return

    now = datetime.now(ZoneInfo(TZ))
    today = now.date()
    today_text = today.isoformat()
    old = load(STATE, {})

    if not FORCE and (
        now.hour < HOUR
        or old.get("last_run_date") == today_text
        or old.get("last_attempt_date") == today_text
    ):
        print("CARD MARKET: ikke tid til ny daglig kørsel")
        return

    old["last_attempt_date"] = today_text
    old["last_attempt_at"] = now.isoformat()
    save(STATE, old)

    watch = load(WATCH, {})
    planned = sum(
        len(set_info["cardIds"])
        for game in watch.get("games", [])
        for set_info in game.get("sets", [])
    )
    planned_requests = sum(
        math.ceil(len(set_info["cardIds"]) / 10)
        for game in watch.get("games", [])
        for set_info in game.get("sets", [])
    )
    print(f"CARD MARKET V1.6: {planned} kort / {planned_requests} requests")

    cards, used, remaining, limit = fetch_cards(watch)
    if len(cards) < planned * 0.8:
        raise RuntimeError(f"Kun {len(cards)}/{planned} kort hentet; state gemmes ikke")

    stamp = now.isoformat()
    next_cards, lows = add_history(cards, old, stamp, today)
    ranked_cards = ranked(list(next_cards.values()))

    old_scores = old.get("set_scores", {}) if isinstance(old.get("set_scores", {}), dict) else {}
    summaries = build_set_summaries(ranked_cards, old_scores=old_scores)
    first_scores = not bool(old_scores)
    opportunities = sealed_opportunities(summaries, price_history_products())
    weekly = now.weekday() == 6

    file, ranked_cards = workbook(
        list(next_cards.values()), summaries, opportunities, stamp, used, remaining
    )

    for game in ("POKÉMON", "LORCANA"):
        discord_market_pulse(
            ranked_cards,
            game,
            summaries,
            opportunities,
            first_scores,
            lows,
            used,
            weekly=weekly,
        )

    if weekly:
        for game in ("POKÉMON", "LORCANA"):
            discord_weekly_report(ranked_cards, game, summaries, opportunities)

    post(
        content=(
            "📎 **Card Market Watch · fuld Excel**\n"
            "Top 20 pr. sæt + Heat Score, movers, Low/Trend/1d/7d/30d, foil-data, "
            "observeret historik og DK sealed-opportunities."
        ),
        file=file,
    )

    set_history = update_set_history(old, summaries, today)
    save(
        STATE,
        {
            "version": 2,
            "last_attempt_date": today_text,
            "last_attempt_at": old.get("last_attempt_at"),
            "last_run_date": today_text,
            "last_run_at": stamp,
            "requests_used": used,
            "rate_limit": limit,
            "rate_remaining": remaining,
            "set_scores": current_score_state(summaries),
            "set_history": set_history,
            "cards": next_cards,
        },
    )
    print(
        f"CARD MARKET færdig: {len(cards)} kort | {used} requests | "
        f"remaining={remaining} | opportunities={len(opportunities)}"
    )


if __name__ == "__main__":
    main()
